from time import strftime, gmtime

import openpyxl
import xlsxwriter
from aiogram import types
from aiogram.dispatcher import FSMContext

from loader import dp, bot
from utils.db_api.users_sql import Database_User

db_user = Database_User()


# @dp.message_handler(text="/btkberish")
# async def reg_uchun(message: types.Message, state: FSMContext):
#     try:
#         await db_user.create()
#     except:
#         pas = 1
#     try:
#         await db_user.create_table_users()
#     except:
#         pass
#     wb_z_u = openpyxl.load_workbook('data/Exel/regold.xlsx')
#     ws_z_u = wb_z_u.active
#     btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
#     name = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
#     nomer = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
#     qoshim_nomer = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
#     manzil = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
#     time = [ws_z_u.cell(row=i, column=7).value for i in range(1, ws_z_u.max_row + 1)]
#     user_id = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
#     # ws_z_u.max_row

#     print(ws_z_u.max_row)
#     for i in range(1, 59):
#         print(i)
#         print(f"btk", btk[i])
#         print(f"name", name[i])
#         print(f"nomer", nomer[i])
#         print(f"qoshim_nomer", qoshim_nomer[i])
#         print(f"manzil", manzil[i])
#         print(f"time", str(time[i])[:10])
#         print(f"user_id", user_id[i])
#         await db_user.add_user(user_id=str(user_id[i]), BTK_id=str(btk[i]), name=str(name[i]), nomer=str(nomer[i]),
#                               qosh_nomer=str(qoshim_nomer[i]),
#                               address=str(manzil[i]), vaqt=str(time[i])[:10])


@dp.message_handler(text="/hammamijozlar")
async def reg_uchun(message: types.Message, state: FSMContext):
    try:
        await db_user.create()
    except:
        pas = 1
    try:
        await db_user.create_table_users()
    except:
        pas=1

    workbook = xlsxwriter.Workbook(f'data/Exel/Clients.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write('A1', "№")
    worksheet.write('B1', "ID raqam")
    worksheet.write('C1', "Ism Familya")
    worksheet.write('D1', "Nomer 1")
    worksheet.write('E1', "Nomer 2")
    worksheet.write('F1', "Viloyat")
    worksheet.write('G1', "Manzil")
    worksheet.write('H1', "Royxatdan otgan sanasi va vaqti")
    worksheet.write('I1', "ID telegram ")

    # try:
    users = await db_user.select_all_users()

    for i in range(len(users)):
            # print(f"user_id", users[i][1])
            # print(f"btk", users[i][2])
            # print(f"name", users[i][3])
            # print(f"nomer", users[i][4])
            # print(f"qoshim_nomer", users[i][5])
            # print(f"qoshim_nomer", users[i][6])
            # print(f"manzil", users[i][7])
            # print(f"time", users[i][8])
    
            # worksheet.write('A1', "№")
            try:
                worksheet.write(f'A{i+2}', users[i][0])
                worksheet.write(f'B{i+2}', f"BTK-{users[i][2]}")
                worksheet.write(f'C{i+2}', users[i][3])
                worksheet.write(f'D{i+2}', users[i][4])
                worksheet.write(f'E{i+2}', users[i][5])
                worksheet.write(f'F{i+2}', users[i][6])
                worksheet.write(f'G{i+2}', users[i][7])
                worksheet.write(f'H{i+2}', users[i][8])
                worksheet.write(f'I{i+2}', users[i][1])
            except:
                pas=1

    # except:
    #     pas=1
    workbook.close()
    with open(f"data/Exel/Clients.xlsx", "rb") as photo_file:
        await bot.send_document(chat_id=message.from_user.id, document=photo_file)


