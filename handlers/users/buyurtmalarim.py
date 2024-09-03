import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import ContentType

from keyboards.default.bosh_menu import buyurtmalar
from keyboards.default.bosh_menu_ru import buyurtmalar_ru
from utils.db_api.users_sql import Database_User

db_user = Database_User()

from loader import bot, dp


class ExelState(StatesGroup):
    exel = State()


@dp.message_handler(text="📋Mening buyurtmalarim")
async def ecr1ho(message: types.Message):
    await message.answer(f"Kerakli bo'limni tanlang.", reply_markup=buyurtmalar)


@dp.message_handler(text="Xitoy omboriga borgan")
async def ecrjychho(message: types.Message):
    try:
        await db_user.create()
    except:
        pas = 1
    user = await db_user.select_user(user_id=str(message.from_user.id))
    btk = user[2]
    wb_z_u = openpyxl.load_workbook('data/Exel/Xitoy_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
    values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0
    summa = 0.0
    soni = 0
    ves = 0.0
    for j in range(1, ws_z_u.max_row):
        if str(btk) == str(values_0_btk[j]):
            test = 1
            sana = str(values_3_sana[j])

            soni = soni + int(values_1_1_soni[j])
            ves = ves + float(values_2_ogirlik[j])
            summa = summa + 7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1])
            await message.answer(f"Trek nomer:  {values_1_trek_nomer[j]}  \n"
                                 f"Og'irlik:     {values_2_ogirlik[j]} kg  \n"
                                 f"Soni:     {values_1_1_soni[j]}   \n"
                                 f"Tarif:      {values_5_tarif[j]}  \n"
                                 f"Summa:      {int(7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1]))}   so'm\n\n"
                                 f"Qabul Vaqti:  {sana[:10]}  \n"
                                 f"Holati:      {values_4_holati[j]}  \n")

    if test == 0:
        await message.answer(f"Sizda hali yuklar mavjud emas\n")
    else:
        await message.answer(f"Umumiy ma'lumot:\n"
                             f"Soni:  {soni}  \n"
                             f"Og'irlik:  {str(ves)[:4]}   kg\n"
                             f"Summa:    {int(summa)}   so'm")


@dp.message_handler(text="Toshkent omboriga kelgan")
async def ecrjychho(message: types.Message):
    try:
        await db_user.create()
    except:
        pas = 1
    user = await db_user.select_user(user_id=str(message.from_user.id))
    btk = user[2]
    wb_z_u = openpyxl.load_workbook('data/Exel/Toshkent_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
    values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0
    summa = 0.0
    soni = 0
    ves = 0.0
    for j in range(1, ws_z_u.max_row):
        if str(btk) == str(values_0_btk[j]):
            test = 1
            sana = str(values_3_sana[j])

            soni = soni + int(values_1_1_soni[j])
            ves = ves + float(values_2_ogirlik[j])
            summa = summa + 7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1])
            await message.answer(f"Trek nomer:  {values_1_trek_nomer[j]}  \n"
                                 f"Og'irlik:     {values_2_ogirlik[j]} kg  \n"
                                 f"Soni:     {values_1_1_soni[j]}   \n"
                                 f"Tarif:      {values_5_tarif[j]}  \n"
                                 f"Summa:      {int(7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1]))}   so'm\n\n"
                                 f"Qabul Vaqti:  {sana[:10]}  \n"
                                 f"Holati:      {values_4_holati[j]}  \n")

    if test == 0:
        await message.answer(f"Sizda hali yuklar mavjud emas\n")
    else:
        await message.answer(f"Umumiy ma'lumot:\n"
                             f"Soni:  {soni}  \n"
                             f"Og'irlik:  {str(ves)[:4]}   kg\n"
                             f"Summa:    {int(summa)}   so'm")





@dp.message_handler(text="📋Мои заказы")
async def ec12ho(message: types.Message):
    await message.answer(f"Выберите нужный раздел.", reply_markup=buyurtmalar_ru)


@dp.message_handler(text="Пришло на склад в Ташкенте")
async def ecrjychho(message: types.Message):
    try:
        await db_user.create()
    except:
        pas = 1
    user = await db_user.select_user(user_id=str(message.from_user.id))
    btk = user[2]
    wb_z_u = openpyxl.load_workbook('data/Exel/Toshkent_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
    values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0
    summa = 0.0
    soni = 0
    ves = 0.0
    j = 0
    for j in range(1, ws_z_u.max_row):
        if str(btk) == str(values_0_btk[j]):
            test = 1
            sana = str(values_3_sana[j])

            soni = soni + int(values_1_1_soni[j])
            ves = ves + float(values_2_ogirlik[j])
            summa = summa + 7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1])

            await message.answer(f"Номер трека: {values_1_trek_nomer[j]}\n"
                                 f"Вес    {values_2_ogirlik[j]} kg\n"
                                 f"Количество:     {values_1_1_soni[j]}   \n"
                                 f"Тариф:      {values_5_tarif[j]}  \n"
                                 f"Сумма:      {int(7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1]))}   сум\n\n"
                                 f"Время приема:  {sana[:10]}  \n"
                                 f"Положение дел:       {values_4_holati[j]}  \n")

    if test == 0:
        await message.answer(f"У вас пока нет грузов\n")
    else:
        await message.answer(f"Общая информация:\n"
                             f"Количество:  {soni}  \n"
                             f"Вес:  {str(ves)[:4]}   кг\n"
                             f"Сумма:  {int(summa)}   сум")


@dp.message_handler(text="Поехал на китайский склад")
async def ecrjychho(message: types.Message):
    try:
        await db_user.create()
    except:
        pas = 1
    user = await db_user.select_user(user_id=str(message.from_user.id))
    btk = user[2]
    wb_z_u = openpyxl.load_workbook('data/Exel/Xitoy_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
    values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0
    summa = 0.0
    soni = 0
    ves = 0.0
    j = 0
    for j in range(1, ws_z_u.max_row):
        if str(btk) == str(values_0_btk[j]):
            test = 1
            sana = str(values_3_sana[j])

            soni = soni + int(values_1_1_soni[j])
            ves = ves + float(values_2_ogirlik[j])
            summa = summa + 7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1])

            await message.answer(f"Номер трека: {values_1_trek_nomer[j]}\n"
                                 f"Вес    {values_2_ogirlik[j]} kg\n"
                                 f"Количество:     {values_1_1_soni[j]}   \n"
                                 f"Тариф:      {values_5_tarif[j]}  \n"
                                 f"Сумма:      {int(7.0 * float(values_2_ogirlik[j]) * float(values_6_kurs[1]))}   сум\n\n"
                                 f"Время приема:  {sana[:10]}  \n"
                                 f"Положение дел:       {values_4_holati[j]}  \n")

    if test == 0:
        await message.answer(f"У вас пока нет грузов\n")
    else:
        await message.answer(f"Общая информация:\n"
                             f"Количество:  {soni}  \n"
                             f"Вес:  {str(ves)[:4]}   кг\n"
                             f"Сумма:  {int(summa)}   сум")
