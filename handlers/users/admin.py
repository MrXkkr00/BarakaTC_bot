import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ContentType

from keyboards.default.common import til
from utils.db_api.users_sql import Database_User

db_user = Database_User()

from loader import dp, bot

admin_menu = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="/exel_xitoy"),
            KeyboardButton(text="/exel_toshkent")
        ],
        [
            # KeyboardButton(text="/Trek_info_yuklash"),
            KeyboardButton(text="/eslatma_toshkent")
        ],
        [
            KeyboardButton(text="/hammamijozlar"),
            # KeyboardButton(text="/trek_raqamlar")
        ],
        [
            KeyboardButton(text="/start"),
        ],
    ],
    resize_keyboard=True
)


@dp.message_handler(text="/adminadmin")
async def admin_panel(message: types.Message):
    await message.answer(f"Admin", reply_markup=admin_menu)


class ExelState_2(StatesGroup):
    exel = State()

class ExelState_t(StatesGroup):
    exel = State()


@dp.message_handler(text="/exel_toshkent")
async def exel(message: types.Message):
    await message.answer(f"Toshkent Exelni yuboring", )
    await ExelState_t.exel.set()


@dp.message_handler(content_types=ContentType.DOCUMENT, state=ExelState_t.exel)
async def exel2(message: types.Message, state: FSMContext):
    file_id_doc = message.document.file_id
    file = await bot.get_file(file_id_doc)
    file_path = file.file_path
    await bot.download_file(file_path, "./data/Exel/Toshkent_yuklar.xlsx")
    await message.answer(f" ToshkentExel yangilandi")
    await state.finish()


@dp.message_handler(text="/exel_xitoy")
async def exel(message: types.Message):
    await message.answer(f"Xitoy Exelni yuboring", )
    await ExelState_2.exel.set()


@dp.message_handler(content_types=ContentType.DOCUMENT, state=ExelState_2.exel)
async def exel2(message: types.Message, state: FSMContext):
    file_id_doc = message.document.file_id
    file = await bot.get_file(file_id_doc)
    file_path = file.file_path
    await bot.download_file(file_path, "./data/Exel/Xitoy_yuklar.xlsx")
    await message.answer(f"Xitoy Exel yangilandi")
    await state.finish()



# @dp.message_handler(text="/eslatma_toshkent")
# async def eslatma(message: types.Message):
#     try:
#         await db_user.create()
#     except:
#         pas = 1
#
#     wb_z_u = openpyxl.load_workbook('data/Exel/Toshkent_yuklar.xlsx')
#     ws_z_u = wb_z_u.active
#     values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
#     values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
#     values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
#     values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
#     values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
#     values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
#     values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
#     values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]
#     wb_z_u.close()
#
#     for i in range(1, ws_z_u.max_row):
#         btk = values_0_btk[i]
#         if not btk is None:
#             user = await db_user.select_user(BTK_id=str(btk))
#             user_id = user[1]
#             sana = str(values_3_sana[i])
#             await bot.send_message(chat_id=user_id, text=f"Trek nomer:  {values_1_trek_nomer[i]}  \n"
#                                                              f"Og'irligi:     {values_2_ogirlik[i]} kg  \n"
#                                                              f"Soni:     {values_1_1_soni[i]}   \n"
#                                                              f"Tarif:      {values_5_tarif[i]}  \n"
#                                                              f"Summa:      {int(7.0 * float(values_2_ogirlik[i]) * float(values_6_kurs[1]))}   so'm\n\n "
#                                                              f"Qabul Vaqti:  {sana[:10]}  \n"
#                                                              f"Holati:      {values_4_holati[i]}  \n")
#             await message.answer(f"{btk}")
#     try:
#         await db_user.dicsconnect()
#     except:
#         pas = 1
#     await message.answer(f"Buyurtmalari haqida xabar barcha mijozlarga yuborildi")



# @dp.message_handler(text="/Trek_info_yuklash")
# async def eslatma(message: types.Message):
#     try:
#         await db_user.create()
#     except:
#         pas = 1
#     wb_z_u = openpyxl.load_workbook('data/Exel/Trek_info.xlsx')
#     ws_z_u = wb_z_u.active
#     values_0_btk = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
#     values_1_1_soni = [ws_z_u.cell(row=i, column=8).value for i in range(1, ws_z_u.max_row + 1)]
#     values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
#     values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
#     values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
#     values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
#     values_5_tarif = [ws_z_u.cell(row=i, column=6).value for i in range(1, ws_z_u.max_row + 1)]
#     values_6_kurs = [ws_z_u.cell(row=i, column=9).value for i in range(1, ws_z_u.max_row + 1)]
#
#     for i in range(1, ws_z_u.max_row):
#         try:
#             btk = values_0_btk[i]
#             print(btk)
#             user = await db_user.select_user(BTK_id=str(btk))
#             user_id = user[1]
#             sana = str(values_3_sana[i])
#             await bot.send_message(chat_id=user_id, text=f"Trek nomer:  {values_1_trek_nomer[i]}  \n"
#                                                          f"Og'irligi:     {values_2_ogirlik[i]} kg  \n"
#                                                          f"Soni:     {values_1_1_soni[i]}   \n"
#                                                          f"Tarif:      {values_5_tarif[i]}  \n"
#                                                          f"Summa:      {int(7.0 * float(values_2_ogirlik[i]) * float(values_6_kurs[1]))}   so'm\n\n "
#                                                          f"Qabul Vaqti:  {sana[:10]}  \n"
#                                                          f"Holati:      {values_4_holati[i]}  \n")
#         except:
#             pas = 1
#
#     await message.answer(f"Buyurtmalari haqida xabar barcha mijozlarga yuborildi")
#


class ExelState_trekt(StatesGroup):
    exel = State()


@dp.message_handler(text="/Trek_info_yuklash")
async def exel(message: types.Message):
    await message.answer(f"Toshkent Exelni yuboring", )
    await ExelState_trekt.exel.set()


@dp.message_handler(content_types=ContentType.DOCUMENT, state=ExelState_trekt.exel)
async def exel2(message: types.Message, state: FSMContext):
    file_id_doc = message.document.file_id
    file = await bot.get_file(file_id_doc)
    file_path = file.file_path
    await bot.download_file(file_path, "./data/Exel/Trek_info.xlsx")
    await message.answer(f" TrekInfoYuklar yangilandi")
    await state.finish()


@dp.message_handler(text='/start', state=[ExelState_trekt.exel, ExelState_2.exel, ExelState_t.exel])
async def stafsdfrt(message: types.Message, state:FSMContext):
    await message.answer(f"Assalomu Alaykum!", reply_markup=til)
    await state.finish()
