import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup

from keyboards.default.bosh_menu import only_boshmenu, only_boshmenu_trek, bosh_menu
from keyboards.default.bosh_menu_ru import only_boshmenu_ru_trek, bosh_menu_ru
from keyboards.default.common import til
from loader import dp



class Trek_Info_yuk(StatesGroup):
    trek_nomer = State()


class Trek_Info_yuk_ru(StatesGroup):
    trek_nomer = State()


@dp.message_handler(text="Вернуться в главное меню",state=Trek_Info_yuk_ru.trek_nomer)
async def bosh_menus324hu(message: types.Message, state:FSMContext):
    await message.answer(f"📲Выберите нужный раздел", reply_markup=bosh_menu_ru)
    await state.finish()


@dp.message_handler(text="Bosh menuga qaytish", state=Trek_Info_yuk.trek_nomer)
async def bosh_safamenus324hu(message: types.Message, state: FSMContext):
    await message.answer(f"📲 Kerakli bo'limni tanlang", reply_markup=bosh_menu)
    await state.finish()


@dp.message_handler(text="/start",state=Trek_Info_yuk_ru.trek_nomer)
async def bosh_mfdsf324hu(message: types.Message, state:FSMContext):
    await message.answer(f"📲Выберите нужный раздел", reply_markup=til)
    await state.finish()


@dp.message_handler(text="/start", state=Trek_Info_yuk.trek_nomer)
async def bosh_safasdfs324hu(message: types.Message, state: FSMContext):
    await message.answer(f"Assalomu Alaykum!", reply_markup=til)
    await state.finish()
    
    

    
    

@dp.message_handler(text="🚚Информация о треке")
async def ecrjychho(message: types.Message, state: FSMContext):
    video = open("./data/video_baraka.mp4", "rb")
    await message.answer_video(video=video,
                               caption=f"✍️Получите информацию о грузе, отправившемся на китайский склад, введя номер ТРЭК.")

    await message.answer(f"Введите номер трека.", reply_markup=only_boshmenu_ru_trek)
    await Trek_Info_yuk_ru.trek_nomer.set()


@dp.message_handler(state=Trek_Info_yuk_ru.trek_nomer)
async def ecr1ho(message: types.Message, state: FSMContext):
    trek_nomer = str(message.text)
    wb_z_u = openpyxl.load_workbook('data/Exel/Xitoy_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_cargo_turi = [ws_z_u.cell(row=i, column=10).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0

    for j in range(1, ws_z_u.max_row):
        if str(trek_nomer) in str(values_1_trek_nomer[j]):
            test = 1
            sana = str(values_3_sana[j])

            await message.answer(f"Номер трека: {values_1_trek_nomer[j]}\n"
                                 f"Cargo turi:  {values_5_cargo_turi[j]}  \n"
                                 f"Вес    {values_2_ogirlik[j]} kg\n"
                                 f"Время приема:  {sana[:10]}  \n"
                                 f"Положение дел:       {values_4_holati[j]}  \n", reply_markup=only_boshmenu_ru_trek)

    if test == 0:
        await message.answer(f"У вас пока нет грузов\n", reply_markup=only_boshmenu_ru_trek)





@dp.message_handler(text="🚚Trek haqida ma'lumot")
async def ecrjych3424ho(message: types.Message, state: FSMContext):
    video = open("./data/video_baraka.mp4", "rb")
    await message.answer_video(video=video,
                               caption=f"✍️TREK raqam kiritish orqali Xitoy omboriga borgan yuklar haqida maʼlumot olish")

    await message.answer(f"Trek raqam kiriting.", reply_markup=only_boshmenu_trek)
    await Trek_Info_yuk.trek_nomer.set()


@dp.message_handler(state=Trek_Info_yuk.trek_nomer)
async def ecrjyc4234hho(message: types.Message, state: FSMContext):
    trek_nomer = str(message.text)
    wb_z_u = openpyxl.load_workbook('data/Exel/Xitoy_yuklar.xlsx')
    ws_z_u = wb_z_u.active
    values_1_trek_nomer = [ws_z_u.cell(row=i, column=2).value for i in range(1, ws_z_u.max_row + 1)]
    values_2_ogirlik = [ws_z_u.cell(row=i, column=3).value for i in range(1, ws_z_u.max_row + 1)]
    values_3_sana = [ws_z_u.cell(row=i, column=4).value for i in range(1, ws_z_u.max_row + 1)]
    values_4_holati = [ws_z_u.cell(row=i, column=5).value for i in range(1, ws_z_u.max_row + 1)]
    values_5_cargo_turi = [ws_z_u.cell(row=i, column=10).value for i in range(1, ws_z_u.max_row + 1)]

    test = 0

    for j in range(1, ws_z_u.max_row):
        if str(trek_nomer) in str(values_1_trek_nomer[j]):
            test = 1
            sana = str(values_3_sana[j])

            await message.answer(f"Trek nomer:  {values_1_trek_nomer[j]}  \n"
                                 f"Cargo turi:  {values_5_cargo_turi[j]}  \n"
                                 f"Og'irlik:     {values_2_ogirlik[j]} kg  \n"
                                 f"Qabul Vaqti:  {sana[:10]}  \n"
                                 f"Holati:      {values_4_holati[j]}  \n", reply_markup=only_boshmenu_trek)

    if test == 0:
        await message.answer(f"Sizda hali yuklar mavjud emas\n", reply_markup=only_boshmenu_trek)



